import os
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

numero_sequencial = "nome do pdf"

# Função para gerar o nome do arquivo baseado no ano atual
def nome_arquivo_excel(ano):
    return f"planilhas_mensais_{ano}.xlsx"

# Função para obter o nome do mês atual em português
def obter_nome_mes(mes_numero):
    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    return meses[mes_numero - 1]

# Função para formatar a aba com títulos e estilos
def formatar_aba(sheet):
    # Definindo os títulos das colunas
    titulos = ["ID", "NOME DO CLIENTE", "CPF/CNPJ", "TELEFONE",
               "DATA/HORA DA COMPRA", "ITENS", "CATEGORIA",
               "FORMA DE PAGAMENTO", "MODALIDADE DE VENDA", "FORMA DE PAGAMENTO"]

    # Aplicando títulos e estilos nas células
    for col, titulo in enumerate(titulos, start=1):  # start=1 para começar da coluna A
        cell = sheet.cell(row=1, column=col)
        cell.value = titulo
        cell.font = Font(name='Times New Roman', size=12, bold=True, italic=False, color="FF0000")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Aplicando bordas
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

        # Ajustando a largura da coluna
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 20

    # Ajustando a altura da linha 1
    sheet.row_dimensions[1].height = 30

# Função para encontrar a próxima linha vazia
def encontrar_proxima_linha(sheet):
    for row in range(2, sheet.max_row + 2):  # Começando da linha 2
        if all(sheet.cell(row=row, column=col).value is None for col in range(1, 11)):  # Colunas de 1 a 10
            return row
    return sheet.max_row + 1  # Se não encontrou, retorna a próxima linha

# Função principal
def main():
    # Obtém o mês e ano atuais
    mes_atual = obter_nome_mes(datetime.now().month)
    ano_atual = datetime.now().year

    # Define o nome do arquivo com base no ano atual
    arquivo_excel = nome_arquivo_excel(ano_atual)

    # Criação da pasta onde o arquivo será salvo
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    nome_pasta = "Planilhas de Controle"
    pasta_destino = os.path.join(desktop_path, nome_pasta)

    # Certifique-se de que o diretório existe, se não, crie-o
    os.makedirs(pasta_destino, exist_ok=True)

    # Caminho completo para o arquivo Excel na pasta criada
    caminho_completo_excel = os.path.join(pasta_destino, arquivo_excel)

    # Verifica se o arquivo Excel para o ano atual já existe
    if not os.path.exists(caminho_completo_excel):
        # Se não existir, cria um novo workbook com as abas dos meses
        workbook = Workbook()

        # Meses do ano
        meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

        # Adiciona uma aba para cada mês e formata
        for mes in meses:
            sheet = workbook.create_sheet(title=mes)
            formatar_aba(sheet)  # Formata a aba

        # Remove a planilha padrão criada automaticamente
        if 'Sheet' in workbook.sheetnames:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)

        # Salva o novo arquivo para o ano atual
        workbook.save(caminho_completo_excel)

    # Carrega o workbook existente
    workbook = openpyxl.load_workbook(caminho_completo_excel)

    # Preenche as informações na aba correspondente ao mês atual
    if mes_atual in workbook.sheetnames:
        sheet_atual = workbook[mes_atual]
        # Encontrar a próxima linha vazia
        proxima_linha = encontrar_proxima_linha(sheet_atual)

        # Preenchendo as informações na próxima linha vazia
        sheet_atual[f'A{proxima_linha}'] = numero_sequencial
        sheet_atual[f'B{proxima_linha}'] = 12345
        sheet_atual[f'C{proxima_linha}'] = 123
        sheet_atual[f'D{proxima_linha}'] = 123
        sheet_atual[f'E{proxima_linha}'] = 123
        sheet_atual[f'F{proxima_linha}'] = 123
        sheet_atual[f'G{proxima_linha}'] = 123
        sheet_atual[f'H{proxima_linha}'] = 123
        sheet_atual[f'I{proxima_linha}'] = 123
        sheet_atual[f'J{proxima_linha}'] = 123

        # Salva as alterações
        workbook.save(caminho_completo_excel)
        print(f"Dados adicionados à aba do mês {mes_atual} na linha {proxima_linha}.")

if __name__ == "__main__":
    main()


## FORMATA TODAS AS ABAS COLOCANDO OS TITULOS, O CONTEÚDO SÓ É COLOCADO NA ABA DO MÊS DECORRENTE UMA INFO ABAIXO DA OUTRA