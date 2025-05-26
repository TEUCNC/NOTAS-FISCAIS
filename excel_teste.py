##OK

import os
import openpyxl
from datetime import datetime
from openpyxl import Workbook

# Função para gerar o nome do arquivo baseado no ano atual
def nome_arquivo_excel(ano):
    return f"planilhas_mensais_{ano}.xlsx"

# Obtém o ano e o mês atuais
ano_atual = datetime.now().year
mes_atual = datetime.now().month

# Define o nome do arquivo com base no ano atual
arquivo_excel = nome_arquivo_excel(ano_atual)

# Criação da pasta onde o arquivo será salvo
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
nome_pasta = "Planilhas de Controle"
pasta_destino = os.path.join(desktop_path, nome_pasta)

# Certifique-se de que o diretório existe, se não, crie-o
os.makedirs(pasta_destino, exist_ok=True)

# Caminho completo para o arquivo Excel na pasta criada
caminho_completo_excel = os.path.join(pasta_destino, arquivo_excel)

# Verifica se o arquivo Excel para o ano atual já existe
if not os.path.exists(caminho_completo_excel):
    # Se não existir, cria um novo workbook com as abas dos meses
    workbook = openpyxl.Workbook()

    # Meses do ano
    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

    # Adiciona uma aba para cada mês
    for mes in meses:
        workbook.create_sheet(title=mes)

    # Remove a planilha padrão criada automaticamente
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # Salva o novo arquivo na pasta criada
    workbook.save(caminho_completo_excel)

# Recarrega o arquivo para editar
workbook = openpyxl.load_workbook(caminho_completo_excel)

# Exemplo de como acessar e manipular cada aba

# Acessando a aba 'Janeiro'
aba_janeiro = workbook['Janeiro']
aba_janeiro['A1'] = "Dados de Janeiro"
aba_janeiro['B1'] = 100

# Acessando a aba 'Fevereiro'
aba_fevereiro = workbook['Fevereiro']
aba_fevereiro['A1'] = "Dados de Fevereiro"
aba_fevereiro['B1'] = 200

# Acessando a aba 'Março'
aba_marco = workbook['Março']
aba_marco['A1'] = "Dados de Março"
aba_marco['B1'] = 300

# Salvando as alterações
workbook.save(caminho_completo_excel)
